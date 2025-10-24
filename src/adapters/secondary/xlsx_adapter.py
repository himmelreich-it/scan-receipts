"""XLSX operations adapter implementation."""

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from ports.xlsx import XLSXPort

logger = logging.getLogger(__name__)


class XLSXAdapter(XLSXPort):
    """XLSX operations implementation using openpyxl."""

    def get_last_sequential_number(self, xlsx_path: Path) -> int:
        """Get the last sequential number from XLSX file.

        Reads from column A starting at row 11 to find the highest number.
        Returns 0 if file is empty or doesn't exist.

        Args:
            xlsx_path: Path to XLSX file.

        Returns:
            Last sequential number, or 0 if none found.
        """
        if not xlsx_path.exists():
            logger.info(f"XLSX file does not exist: {xlsx_path}")
            return 0

        try:
            workbook = load_workbook(xlsx_path)
            sheet = workbook.active
            if sheet is None:
                raise ValueError("No active sheet found in workbook")

            # Find the last used row starting from row 11
            last_number = 0
            for row in sheet.iter_rows(min_row=11, max_col=1, values_only=True):
                cell_value = row[0]
                if cell_value is not None and isinstance(cell_value, (int, float)):
                    last_number = max(last_number, int(cell_value))

            workbook.close()
            logger.info(f"Last sequential number in XLSX: {last_number}")
            return last_number

        except Exception as e:
            logger.error(f"Failed to read last sequential number from XLSX: {e}")
            return 0

    def append_entries(
        self, xlsx_path: Path, entries: List[Dict[str, str]]
    ) -> List[int]:
        """Append receipt entries to XLSX file.

        Appends entries starting from first empty row (found by checking Column B).
        Reads sequential numbers from Column A (already filled in spreadsheet).
        Writes to: B=Date, F=Description, J=Amount, K=VAT, S=Notes (NOT Column A).

        Args:
            xlsx_path: Path to XLSX file.
            entries: List of receipt entries with keys: date, description, amount, tax, currency

        Returns:
            List of sequential numbers used for each entry (read from Column A).

        Raises:
            OSError: If file operations fail.
            ValueError: If no sequential number found in Column A.
        """
        if not xlsx_path.exists():
            raise OSError(f"XLSX file does not exist: {xlsx_path}")

        try:
            workbook = load_workbook(xlsx_path)
            sheet = workbook.active
            if sheet is None:
                raise ValueError("No active sheet found in workbook")

            # Find the next empty row starting from row 11
            # Check Column B (date column) as it's the reliable indicator of filled rows
            next_row = 11
            while True:
                cell_value = sheet.cell(row=next_row, column=2).value  # Column B
                if cell_value is None or str(cell_value).strip() == "":
                    break
                next_row += 1

            # Append each entry
            # Read sequential numbers from Column A (already filled in spreadsheet)
            used_numbers = []
            for entry in entries:
                # Read sequential number from Column A (don't write it)
                sequential_number = sheet.cell(row=next_row, column=1).value
                if sequential_number is None or not isinstance(sequential_number, (int, float)):
                    raise ValueError(
                        f"No sequential number found in Column A at row {next_row}"
                    )

                current_number = int(sequential_number)
                used_numbers.append(current_number)

                # Column B: Date (convert from dd-MM-YYYY to d-MMM-yy format)
                date_str = entry.get("date", "")
                formatted_date = self._format_date_for_xlsx(date_str)
                sheet.cell(row=next_row, column=2).value = formatted_date

                # Column F: Description
                sheet.cell(row=next_row, column=6).value = entry.get("description", "")

                # Column J: Total amount
                amount_str = entry.get("amount", "0")
                try:
                    amount = float(amount_str)
                    sheet.cell(row=next_row, column=10).value = amount
                except ValueError:
                    sheet.cell(row=next_row, column=10).value = 0

                # Column K: VAT amount
                tax_str = entry.get("tax", "0")
                try:
                    tax = float(tax_str)
                    sheet.cell(row=next_row, column=11).value = tax
                except ValueError:
                    sheet.cell(row=next_row, column=11).value = 0

                # Column S: Notes (for non-EUR currencies)
                currency = entry.get("currency", "EUR")
                if currency != "EUR":
                    notes = f"{currency} {amount_str}"
                    sheet.cell(row=next_row, column=19).value = notes

                next_row += 1

            # Save workbook
            workbook.save(xlsx_path)
            workbook.close()

            logger.info(
                f"Appended {len(entries)} entries to XLSX using numbers {used_numbers[0]}-{used_numbers[-1]}"
            )
            return used_numbers

        except Exception as e:
            logger.error(f"Failed to append entries to XLSX: {e}")
            raise OSError(f"Failed to append entries to XLSX: {e}")

    def xlsx_file_exists(self, xlsx_path: Path) -> bool:
        """Check if XLSX file exists.

        Args:
            xlsx_path: Path to XLSX file.

        Returns:
            True if file exists, False otherwise.
        """
        return xlsx_path.exists()

    def _format_date_for_xlsx(self, date_str: str) -> date:
        """Convert date from dd-MM-YYYY to datetime.date object for Excel.

        Excel will format the date object according to the cell's number format.
        This prevents Excel from treating the date as text (with apostrophe prefix).

        Args:
            date_str: Date string in dd-MM-YYYY format.

        Returns:
            datetime.date object that Excel will recognize and format as a date.

        Examples:
            >>> _format_date_for_xlsx("10-07-2025")
            datetime.date(2025, 7, 10)
            >>> _format_date_for_xlsx("01-01-2025")
            datetime.date(2025, 1, 1)
        """
        try:
            date_obj = datetime.strptime(date_str, "%d-%m-%Y")
            # Return date object, not string - Excel will format it properly
            return date_obj.date()
        except ValueError:
            logger.warning(f"Failed to parse date: {date_str}, using today's date")
            return datetime.now().date()
